import openpyxl
path = "C:/Users/bobbu.yaswanth/Desktop/OpenPyXl.xlsx"
workbook = openpyxl.load_workbook(path)
sheet = workbook.active  #workbook.get_sheet_by_name("Sheet1")

rows = sheet.max_row
cols = sheet.max_column
#print(rows)
#print(cols)

for r in range(1,rows+1):
    for c in range(1,cols+1):
        sheet.cell(row=r,column=c).value = "Welcome"

workbook.save(path)