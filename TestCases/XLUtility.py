import openpyxl


def getrowcount(file,sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetName)
    return(sheet.max_row)
def getcolumncount(file,sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetName)
    return(sheet.max_column)
def readdata(file,sheetName,rownum,columnnum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetName)
    return sheet.cell(row=rownum,column=columnnum).value
def writedata(file,sheetName,rownum,columnnum,data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetName)
    sheet.cell(row=rownum,column=columnnum).value = data
    workbook.save(file)